"""
Excel Export Service for generating downloadable Excel files
"""

import io
import os
import sys
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta
import pandas as pd

# Add utils to path
utils_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../../utils'))
if utils_path not in sys.path:
    sys.path.append(utils_path)

from utils.logger import setup_logger

logger = setup_logger(__name__)


class ExcelExportService:
    """Service for generating Excel files from data"""
    
    def __init__(self):
        self.default_sheet_name = "Data Export"
    
    def create_excel_file(
        self,
        data: List[Dict[str, Any]],
        columns: List[Dict[str, str]],
        sheet_name: Optional[str] = None,
        title: Optional[str] = None
    ) -> bytes:
        """
        Create Excel file from data with proper formatting
        
        Args:
            data: List of dictionaries containing the data
            columns: List of dictionaries with 'key' and 'header' for each column
            sheet_name: Name of the Excel sheet
            title: Optional title for the Excel sheet
            
        Returns:
            bytes: Excel file content as bytes
        """
        try:
            logger.info(f"Creating Excel file with {len(data)} records and {len(columns)} columns")
            
            if not data:
                logger.warning("No data provided, creating empty Excel with headers only")
                # Create empty DataFrame with just headers
                headers = [col['header'] for col in columns]
                df = pd.DataFrame(columns=headers)
            else:
                # Create DataFrame from data
                processed_data = []
                for index, row in enumerate(data, 1):
                    processed_row = {'No': index}  # Add row number
                    for col in columns:
                        if col['key'] != 'no':  # Skip 'no' as we handle it separately
                            processed_row[col['header']] = row.get(col['key'], '')
                    processed_data.append(processed_row)
                
                df = pd.DataFrame(processed_data)
            
            # Create Excel file in memory
            buffer = io.BytesIO()
            
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                sheet_name = sheet_name or self.default_sheet_name
                
                # Write DataFrame to Excel
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
                
                # Get the workbook and worksheet to apply formatting
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Apply formatting
                self._apply_excel_formatting(worksheet, df, title)
            
            buffer.seek(0)
            excel_bytes = buffer.getvalue()
            buffer.close()
            
            logger.info(f"Excel file created successfully, size: {len(excel_bytes)} bytes")
            return excel_bytes
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {str(e)}")
            raise Exception(f"Failed to create Excel file: {str(e)}")
    
    def _apply_excel_formatting(self, worksheet, df: pd.DataFrame, title: Optional[str] = None):
        """
        Apply formatting to Excel worksheet
        
        Args:
            worksheet: openpyxl worksheet object
            df: pandas DataFrame
            title: Optional title for the sheet
        """
        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Border styling
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply header formatting
            for col_num, _ in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Apply borders to data cells
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = thin_border
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set minimum and maximum column width
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add title if provided
            if title:
                worksheet.insert_rows(1)
                title_cell = worksheet.cell(row=1, column=1, value=title)
                title_cell.font = Font(bold=True, size=14)
                title_cell.alignment = Alignment(horizontal="center")
                
                # Merge cells for title
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            
            logger.debug("Excel formatting applied successfully")
            
        except Exception as e:
            logger.warning(f"Error applying Excel formatting: {str(e)}")
            # Continue without formatting if there's an error
    
    def generate_filename(
        self,
        export_type: str,
        dealer_id: str,
        date_from: str,
        date_to: str
    ) -> str:
        """
        Generate standardized filename for Excel export
        
        Args:
            export_type: Type of export (e.g., 'WorkOrder', 'NJB_NSC', 'HLO')
            dealer_id: Dealer ID
            date_from: Start date
            date_to: End date
            
        Returns:
            str: Generated filename
        """
        # Convert UTC to Indonesian time (WIB = GMT+7)
        indonesia_time = datetime.now() + timedelta(hours=7)
        timestamp = indonesia_time.strftime("%Y%m%d_%H%M%S")
        filename = f"{export_type}_Export_{dealer_id}_{date_from.replace('-', '')}_{date_to.replace('-', '')}_{timestamp}.xlsx"
        return filename